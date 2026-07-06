"""
Lead Filter System
-------------------
Filters a leads table in Postgres/Supabase and exports qualified results to .xlsx.

Reads DATABASE_URL (or SUPABASE_DB_URL) from a .env file in the working directory.
"""

import os
import re
import sys

import psycopg2
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from rich.console import Console
from rich.panel import Panel
from rich.prompt import Confirm, Prompt
from rich.table import Table

console = Console()

# --------------------------------------------------------------------------
# Config — reads your Supabase connection string from .env
# --------------------------------------------------------------------------

load_dotenv()  # looks for a .env file in the current working directory

DATABASE_URL = os.environ.get("DATABASE_URL") or os.environ.get("SUPABASE_DB_URL")

if not DATABASE_URL:
    console.print(
        "[bold red]No DATABASE_URL (or SUPABASE_DB_URL) found in your .env file.[/bold red]\n"
        "[dim]Add a line like: DATABASE_URL=postgresql://postgres:PASSWORD@HOST:5432/postgres[/dim]"
    )
    sys.exit(1)

IDENTIFIER_RE = re.compile(r"^[a-zA-Z_][a-zA-Z0-9_]*$")

SORT_OPTIONS = {
    "1": ("Reviews (high to low)", "reviews DESC, rating DESC, name ASC"),
    "2": ("Rating (high to low)", "rating DESC, reviews DESC, name ASC"),
    "3": ("Name (A to Z)", "name ASC"),
}


# --------------------------------------------------------------------------
# Input helpers
# --------------------------------------------------------------------------

def ask_number(label, cast, allow_blank=True):
    """Prompt until the input is blank (if allowed) or a valid number."""
    while True:
        raw = Prompt.ask(f"[bold]{label}[/bold]", default="").strip()
        if not raw:
            if allow_blank:
                return None
            console.print("[red]A value is required.[/red]")
            continue
        try:
            return cast(raw)
        except ValueError:
            console.print(f"[red]'{raw}' isn't a valid number. Try again.[/red]")


def ask_identifier(label, default):
    """Prompt for a SQL identifier (table name) and validate it strictly."""
    while True:
        raw = Prompt.ask(f"[bold]{label}[/bold]", default=default).strip().lower()
        if IDENTIFIER_RE.match(raw):
            return raw
        console.print(
            "[red]Table names can only contain letters, numbers, and "
            "underscores, and can't start with a number.[/red]"
        )


# --------------------------------------------------------------------------
# UI: intro
# --------------------------------------------------------------------------

console.print()
console.print(
    Panel.fit(
        "[bold cyan]LEAD FILTER SYSTEM[/bold cyan]\n"
        "[dim]Filter your database and export qualified leads[/dim]",
        border_style="cyan",
    )
)

table_name = ask_identifier("Database Table", "dentists_mumbai")

console.print(
    "\n[dim]Website filter — 1: has a website   2: no website   3: either[/dim]"
)
website_filter = Prompt.ask(
    "[bold]Website[/bold]", choices=["1", "2", "3"], default="3"
)

min_rating = ask_number("Minimum Rating (0-5, blank = no limit)", float)
max_rating = ask_number("Maximum Rating (0-5, blank = no limit)", float)
if min_rating is not None and max_rating is not None and min_rating > max_rating:
    console.print("[yellow]Minimum rating was higher than maximum — swapping them.[/yellow]")
    min_rating, max_rating = max_rating, min_rating

min_reviews = ask_number("Minimum Reviews (blank = no limit)", int)
max_reviews = ask_number("Maximum Reviews (blank = no limit)", int)
if min_reviews is not None and max_reviews is not None and min_reviews > max_reviews:
    console.print("[yellow]Minimum reviews was higher than maximum — swapping them.[/yellow]")
    min_reviews, max_reviews = max_reviews, min_reviews

city = Prompt.ask("[bold]City Contains[/bold]", default="").strip()

sort_options_str = ", ".join(f"{k}: {v[0]}" for k, v in SORT_OPTIONS.items())
console.print(f"\n[dim]Sort order — {sort_options_str}[/dim]")
sort_choice = Prompt.ask("[bold]Sort By[/bold]", choices=list(SORT_OPTIONS), default="1")

limit_raw = ask_number("Limit results (blank = no limit)", int)

output_file = Prompt.ask("[bold]Output Filename[/bold]", default="filtered_leads").strip()
if not output_file.lower().endswith(".xlsx"):
    output_file += ".xlsx"

# --------------------------------------------------------------------------
# Build query (table name validated above; every value is parameterized)
# --------------------------------------------------------------------------

query = f"""
SELECT name, phone, address, rating, reviews, website
FROM {table_name}
WHERE TRUE
"""
params = []

if website_filter == "1":
    query += " AND website IS NOT NULL AND TRIM(website) <> ''"
elif website_filter == "2":
    query += " AND (website IS NULL OR TRIM(website) = '')"

if min_rating is not None:
    query += " AND COALESCE(rating, 0) >= %s"
    params.append(min_rating)
if max_rating is not None:
    query += " AND COALESCE(rating, 0) <= %s"
    params.append(max_rating)
if min_reviews is not None:
    query += " AND COALESCE(reviews, 0) >= %s"
    params.append(min_reviews)
if max_reviews is not None:
    query += " AND COALESCE(reviews, 0) <= %s"
    params.append(max_reviews)
if city:
    query += " AND address ILIKE %s"
    params.append(f"%{city}%")

query += f" ORDER BY {SORT_OPTIONS[sort_choice][1]}"

if limit_raw is not None:
    query += " LIMIT %s"
    params.append(limit_raw)

# --------------------------------------------------------------------------
# Summary + confirmation
# --------------------------------------------------------------------------

summary = Table(show_header=False, box=None, padding=(0, 1))
summary.add_row("Table", table_name)
summary.add_row("Website", {"1": "Has website", "2": "No website", "3": "Either"}[website_filter])
summary.add_row("Rating range", f"{min_rating or 0} – {max_rating or 5}")
summary.add_row("Reviews range", f"{min_reviews or 0} – {max_reviews or '∞'}")
summary.add_row("City contains", city or "(any)")
summary.add_row("Sort by", SORT_OPTIONS[sort_choice][0])
summary.add_row("Limit", str(limit_raw) if limit_raw is not None else "(none)")
summary.add_row("Output file", output_file)

console.print()
console.print(Panel(summary, title="Filter Summary", border_style="green"))

if not Confirm.ask("\nRun this query?", default=True):
    console.print("[yellow]Cancelled.[/yellow]")
    sys.exit(0)

# --------------------------------------------------------------------------
# Execute
# --------------------------------------------------------------------------

try:
    with console.status("[bold green]Connecting to database..."):
        conn = psycopg2.connect(DATABASE_URL, sslmode="require")
except psycopg2.OperationalError as e:
    console.print(f"[bold red]Could not connect to the database:[/bold red] {e}")
    sys.exit(1)

try:
    with conn, conn.cursor() as cur:
        with console.status("[bold green]Running query..."):
            cur.execute(query, params)
            columns = [desc[0] for desc in cur.description]
            rows = cur.fetchall()
except psycopg2.Error as e:
    console.print(f"[bold red]Query failed:[/bold red] {e}")
    console.print(f"[dim]{cur.query.decode() if cur.query else query}[/dim]")
    conn.close()
    sys.exit(1)

conn.close()

if not rows:
    console.print("\n[yellow]No leads matched these filters.[/yellow]")
    sys.exit(0)

# --------------------------------------------------------------------------
# Display a preview
# --------------------------------------------------------------------------

preview_count = min(len(rows), 20)
result_table = Table(title=f"Results (showing {preview_count} of {len(rows)})")
for col in columns:
    result_table.add_column(col.capitalize())

for row in rows[:preview_count]:
    row_dict = dict(zip(columns, row))
    rating_val = row_dict.get("rating")
    rating_str = f"{rating_val:.1f}" if rating_val is not None else "-"
    if rating_val is not None and rating_val >= 4.5:
        rating_str = f"[bold green]{rating_str}[/bold green]"
    elif rating_val is not None and rating_val < 3.5:
        rating_str = f"[bold red]{rating_str}[/bold red]"

    result_table.add_row(*[
        rating_str if col == "rating" else ("" if v is None else str(v))
        for col, v in zip(columns, row)
    ])

console.print()
console.print(result_table)

# --------------------------------------------------------------------------
# Export
# --------------------------------------------------------------------------

wb = Workbook()
ws = wb.active
ws.title = "Leads"

ws.append(columns)
for cell in ws[1]:
    cell.font = Font(bold=True)

for row in rows:
    ws.append(list(row))

# Reasonable auto column widths based on content length
for col_idx, col_name in enumerate(columns, start=1):
    max_len = len(col_name)
    for row in rows:
        val = row[col_idx - 1]
        if val is not None:
            max_len = max(max_len, len(str(val)))
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

wb.save(output_file)

console.print(f"\n[bold green]Exported {len(rows)} leads to {output_file}[/bold green]")