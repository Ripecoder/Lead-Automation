#!/usr/bin/env python3
"""
review_and_collect_leads.py

One merged tool that lets you:
  - Name your output file once, at the start
  - Browse rows in any .xlsx file in the current folder
  - Inspect any row's website by opening it in Chrome (Enter)
  - Check the rows you want to keep (Space)
  - Finish this file (Tab) -> your checked rows are saved immediately
    and you're dropped straight back on the file menu, no interruption
  - Open the next file and repeat — everything accumulates into the
    same output workbook (inside "finalised_leads"), one sheet per
    source file

Note: a bare "Shift" keypress can't be detected by a terminal program —
modifier keys alone send no byte. Tab is used instead as the "I'm done
with this file" key.

Controls in the row browser:
  ↑ / ↓   move
  Enter   open the highlighted row's website in Chrome (inspect)
  Space   toggle selection on the highlighted row
  a / n   select all / select none
  Tab     done with this file -> save & go straight to the file menu
  Esc     back to the file menu (discards unsaved selections for this file)
  Ctrl+C  quit immediately from anywhere

Dependencies: rich, openpyxl  (pip install rich openpyxl)
"""

import os
import platform
import re
import shutil
import subprocess
import sys

try:
    from rich.console import Console, Group
    from rich.live import Live
    from rich.panel import Panel
    from rich.prompt import Prompt
    from rich.table import Table
    from rich.text import Text
except ImportError:
    print("This script needs 'rich'. Install it with:  pip install rich")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
except ImportError:
    print("This script needs 'openpyxl'. Install it with:  pip install openpyxl")
    sys.exit(1)

console = Console()

OUTPUT_FOLDER = "finalised_leads"
URL_REGEX = re.compile(r"(https?://[^\s\"'<>]+|www\.[^\s\"'<>]+)", re.IGNORECASE)


# --------------------------------------------------------------------------
# Cross-platform single-keypress reader
# --------------------------------------------------------------------------

def _get_key():
    """Block for one keypress, returns a normalized token:
    'UP', 'DOWN', 'SPACE', 'ENTER', 'TAB', 'ESC', 'QUIT', or a lowercase letter."""
    if os.name == "nt":
        import msvcrt
        ch = msvcrt.getch()
        if ch in (b"\x00", b"\xe0"):
            ch2 = msvcrt.getch()
            return {b"H": "UP", b"P": "DOWN"}.get(ch2, "")
        if ch == b"\r":
            return "ENTER"
        if ch == b" ":
            return "SPACE"
        if ch == b"\t":
            return "TAB"
        if ch == b"\x1b":
            return "ESC"
        if ch == b"\x03":
            return "QUIT"
        try:
            return ch.decode().lower()
        except UnicodeDecodeError:
            return ""
    else:
        import termios
        import tty
        fd = sys.stdin.fileno()
        old_settings = termios.tcgetattr(fd)
        try:
            tty.setcbreak(fd)
            ch = sys.stdin.read(1)
            if ch == "\x1b":
                import select
                if select.select([sys.stdin], [], [], 0.01)[0]:
                    rest = sys.stdin.read(2)
                    if rest == "[A":
                        return "UP"
                    if rest == "[B":
                        return "DOWN"
                return "ESC"
            if ch in ("\r", "\n"):
                return "ENTER"
            if ch == " ":
                return "SPACE"
            if ch == "\t":
                return "TAB"
            if ch == "\x03":
                return "QUIT"
            return ch.lower()
        finally:
            termios.tcsetattr(fd, termios.TCSADRAIN, old_settings)


# --------------------------------------------------------------------------
# Chrome launching
# --------------------------------------------------------------------------

def find_chrome():
    system = platform.system()
    if system == "Darwin":
        if os.path.exists("/Applications/Google Chrome.app"):
            return ["open", "-a", "Google Chrome"]
        return None
    if system == "Windows":
        candidates = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
        ]
        for path in candidates:
            if os.path.exists(path):
                return [path]
        return None
    for cmd in ("google-chrome", "google-chrome-stable", "chromium-browser", "chromium"):
        if shutil.which(cmd):
            return [cmd]
    return None


_CHROME_CMD = find_chrome()


def open_in_chrome(url):
    if _CHROME_CMD:
        try:
            subprocess.Popen(
                _CHROME_CMD + [url], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
            )
            return True, "chrome"
        except Exception:
            pass
    try:
        import webbrowser
        return webbrowser.open(url, new=2), "default browser"
    except Exception:
        return False, "default browser"


_DOMAIN_LIKE = re.compile(r"^[a-zA-Z0-9-]+(\.[a-zA-Z0-9-]+)+(/[^\s]*)?$")


def normalize_url(raw):
    if not raw:
        return None
    url = str(raw).strip().rstrip(').,;:!?"\'')
    if not url:
        return None
    lower = url.lower()
    if lower.startswith(("http://", "https://")):
        return url
    if lower.startswith("www."):
        return "https://" + url
    # bare domain with no scheme, e.g. "smithdental.com" or "clinic.co.in/home"
    if _DOMAIN_LIKE.match(url):
        return "https://" + url
    return None


# --------------------------------------------------------------------------
# File menu (simple single-select, scrollable)
# --------------------------------------------------------------------------

def select_file_menu(options, title, subtitle=None, max_visible=14):
    if not options:
        return None
    cursor = 0
    offset = 0
    n = len(options)
    window = min(max_visible, n)

    def render():
        nonlocal offset
        if cursor < offset:
            offset = cursor
        elif cursor >= offset + window:
            offset = cursor - window + 1
        table = Table.grid(padding=(0, 1))
        table.add_column()
        if offset > 0:
            table.add_row(Text("  ↑ more above", style="dim"))
        for i in range(offset, min(offset + window, n)):
            pointer = "❯" if i == cursor else " "
            style = "bold cyan" if i == cursor else None
            table.add_row(Text.from_markup(f"{pointer} {options[i]}", style=style))
        if offset + window < n:
            table.add_row(Text("  ↓ more below", style="dim"))
        help_text = "[dim]↑/↓ move   enter select   esc quit[/dim]"
        panel = Panel(table, title=f"[bold]{title}[/bold]", subtitle=subtitle, border_style="cyan")
        return Group(panel, Text.from_markup(help_text))

    with Live(render(), console=console, screen=False, auto_refresh=False) as live:
        while True:
            key = _get_key()
            if key == "UP":
                cursor = (cursor - 1) % n
            elif key == "DOWN":
                cursor = (cursor + 1) % n
            elif key == "ENTER":
                live.stop()
                return cursor
            elif key == "ESC":
                live.stop()
                return None
            elif key == "QUIT":
                live.stop()
                console.print("\n[yellow]Cancelled.[/yellow]")
                sys.exit(0)
            live.update(render(), refresh=True)


# --------------------------------------------------------------------------
# Row browser: inspect (Enter opens Chrome) + select (Space) + finish (Tab)
# --------------------------------------------------------------------------

def browse_and_select_rows(header, rows, url_col_idx, title, max_visible=12):
    """
    Returns a list of selected row indices when the user presses Tab
    ("done with this file"), or None if the user pressed Esc (go back
    to the file menu, discarding any in-progress selections).
    """
    n = len(rows)
    cursor = 0
    offset = 0
    checked = [False] * n
    window = min(max_visible, n)
    status = ""

    def row_text(i):
        parts = [str(v) for v in rows[i] if v is not None and str(v).strip() != ""]
        text = " | ".join(parts) if parts else "(empty row)"
        if len(text) > 100:
            text = text[:99] + "…"
        return text

    def render():
        nonlocal offset
        if cursor < offset:
            offset = cursor
        elif cursor >= offset + window:
            offset = cursor - window + 1
        table = Table.grid(padding=(0, 1))
        table.add_column()
        if offset > 0:
            table.add_row(Text("  ↑ more above", style="dim"))
        for i in range(offset, min(offset + window, n)):
            pointer = "❯" if i == cursor else " "
            box = "[bold green]✔[/bold green]" if checked[i] else " "
            line = f"{pointer} [{box}] [dim]{i + 1}.[/dim] {row_text(i)}"
            style = "bold cyan" if i == cursor else None
            table.add_row(Text.from_markup(line, style=style))
        if offset + window < n:
            table.add_row(Text("  ↓ more below", style="dim"))
        help_text = (
            "[dim]↑/↓ move   enter inspect url in chrome   space select   "
            "a all   n none   tab done with this file   esc back[/dim]"
        )
        selected_count = sum(checked)
        panel = Panel(
            table,
            title=f"[bold]{title}[/bold]",
            subtitle=f"{selected_count} selected · {n} rows",
            border_style="cyan",
        )
        items = [panel]
        if status:
            items.append(Text.from_markup(status))
        items.append(Text.from_markup(help_text))
        return Group(*items)

    with Live(render(), console=console, screen=False, auto_refresh=False) as live:
        while True:
            key = _get_key()
            if key == "UP":
                cursor = (cursor - 1) % n
                status = ""
            elif key == "DOWN":
                cursor = (cursor + 1) % n
                status = ""
            elif key == "SPACE":
                checked[cursor] = not checked[cursor]
                status = ""
            elif key == "a":
                checked = [True] * n
                status = ""
            elif key == "n":
                checked = [False] * n
                status = ""
            elif key == "ENTER":
                if url_col_idx is None:
                    status = "[yellow]No website column found in this file.[/yellow]"
                else:
                    url = normalize_url(rows[cursor][url_col_idx] if url_col_idx < len(rows[cursor]) else None)
                    if not url:
                        status = "[yellow]No valid URL on this row.[/yellow]"
                    else:
                        ok, method = open_in_chrome(url)
                        tag = "Chrome" if method == "chrome" else "default browser"
                        status = (
                            f"[green]Opened in {tag}:[/green] [dim]{url}[/dim]" if ok
                            else f"[red]Failed to open:[/red] [dim]{url}[/dim]"
                        )
            elif key == "TAB":
                live.stop()
                return [i for i, c in enumerate(checked) if c]
            elif key == "ESC":
                live.stop()
                return None
            elif key == "QUIT":
                live.stop()
                console.print("\n[yellow]Cancelled.[/yellow]")
                sys.exit(0)
            live.update(render(), refresh=True)


# --------------------------------------------------------------------------
# Data helpers
# --------------------------------------------------------------------------

def find_xlsx_files(folder="."):
    return sorted(
        f for f in os.listdir(folder)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    )


def read_sheet(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []
    header = [str(c) if c is not None else "" for c in all_rows[0]]
    rows = [list(r) for r in all_rows[1:] if any(v is not None for v in r)]
    return header, rows


def find_url_column(header, rows):
    """Prefer a column literally named website/url/link; otherwise guess by
    scanning a sample of rows for URL-shaped values."""
    lowered = [h.strip().lower() for h in header]
    for candidate in ("website", "url", "link"):
        if candidate in lowered:
            return lowered.index(candidate)

    for col_idx in range(len(header)):
        for row in rows[:20]:
            if col_idx < len(row) and row[col_idx] and URL_REGEX.search(str(row[col_idx])):
                return col_idx
    return None


INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def sanitize_sheet_name(name, existing_titles):
    base = os.path.splitext(name)[0]
    base = INVALID_SHEET_CHARS.sub("_", base).strip() or "Sheet"
    base = base[:31]
    candidate = base
    n = 2
    while candidate in existing_titles:
        suffix = f" ({n})"
        candidate = base[: 31 - len(suffix)] + suffix
        n += 1
    return candidate


def load_or_create_output(path):
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def get_or_create_sheet(wb, source_filename, header):
    """Reuse the sheet for this source file if we've already visited it
    this session; otherwise create a new one with a bold header row."""
    target_title = os.path.splitext(source_filename)[0][:31]
    if target_title in wb.sheetnames:
        return wb[target_title]
    title = sanitize_sheet_name(source_filename, wb.sheetnames)
    ws = wb.create_sheet(title=title)
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    return ws


def count_total_rows(wb):
    return sum(max(ws.max_row - 1, 0) for ws in wb.worksheets)


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main():
    console.print()
    console.print(
        Panel.fit(
            "[bold cyan]REVIEW & COLLECT LEADS[/bold cyan]\n"
            "[dim]Inspect websites in Chrome, check the rows you want, "
            "they land straight in finalised_leads[/dim]",
            border_style="cyan",
        )
    )

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    output_filename = Prompt.ask(
        "\n[bold]Output file name[/bold] [dim](saved inside finalised_leads/, "
        "reused if it already exists)[/dim]",
        default="collected.xlsx",
    ).strip()
    if not output_filename.lower().endswith(".xlsx"):
        output_filename += ".xlsx"
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)
    wb = load_or_create_output(output_path)

    console.print(f"[dim]Everything you select will be saved to:[/dim] [bold]{output_path}[/bold]")

    if _CHROME_CMD is None:
        console.print(
            "[dim]Note: Google Chrome wasn't found on this system — the "
            "'inspect' action will fall back to your default browser.[/dim]"
        )
    console.print()

    files = find_xlsx_files(".")
    if not files:
        console.print("[bold red]No .xlsx files found in the current folder.[/bold red]")
        sys.exit(1)

    # Level 1: file menu. Esc exits the program entirely.
    while True:
        total_so_far = count_total_rows(wb)
        file_idx = select_file_menu(
            files,
            title="Select a file",
            subtitle=f"{len(files)} found · {total_so_far} rows collected so far",
        )
        if file_idx is None:
            console.print(f"\n[yellow]Done. Saved to {output_path}[/yellow]")
            sys.exit(0)

        chosen = files[file_idx]

        with console.status(f"[bold green]Reading {chosen}..."):
            try:
                header, rows = read_sheet(chosen)
            except Exception as e:
                console.print(f"[bold red]Error reading '{chosen}':[/bold red] {e}")
                continue

        if not rows:
            console.print(f"[yellow]No data rows found in '{chosen}'.[/yellow]")
            continue

        url_col_idx = find_url_column(header, rows)

        # Level 2: row browser for this file. Esc goes back to the file menu.
        # Tab saves immediately (no prompt) and also returns to the file menu.
        selected_indices = browse_and_select_rows(header, rows, url_col_idx, title=chosen)

        if selected_indices is None:
            continue  # Esc: back to file menu, nothing saved

        if not selected_indices:
            console.print(f"[yellow]No rows selected from '{chosen}' — nothing saved.[/yellow]\n")
            continue

        selected_rows = [rows[i] for i in selected_indices]
        ws = get_or_create_sheet(wb, chosen, header)
        for row in selected_rows:
            ws.append(row)
        wb.save(output_path)

        console.print(
            f"[bold green]Saved {len(selected_rows)} row(s) from '{chosen}' "
            f"(sheet '{ws.title}').[/bold green]\n"
        )
        # Straight back to the file menu — no prompt, no interruption.


if __name__ == "__main__":
    main()